#!/usr/bin/python3
# -*- coding: utf-8 -*-
# Copyright 2019 Shieber
# All Rights Reserved.
#
#    Licensed under the Apache License, Version 2.0 (the "License"); you may
#    not use this file except in compliance with the License. You may obtain
#    a copy of the License at
#
#         http://www.apache.org/licenses/LICENSE-2.0
#
#    Unless required by applicable law or agreed to in writing, software
#    distributed under the License is distributed on an "AS IS" BASIS, WITHOUT
#    WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the
#    License for the specific language governing permissions and limitations
#    under the License.
#    1.按照每年的消费信息建立统计表 
#    2.按照年月日保存每笔消费信息   
#    3.保存的信息格式为xlsx         
#    4.后期将加入按照每月总消费统计信息的功能      
#    5.后期将加入将统计信息展示为图片的功能         

import os,re,sys,time
import requests
import openpyxl
import datetime
import poplib 

from email.parser import Parser 
from email.utils  import parseaddr 
from email.header import decode_header 

__version__ = '0.1'

class DebtManage():
    def __init__(self):
	'''
	   初始化邮件和excel表的位置，
	   时间和汇率信息等，所有的值
	   依据自己的电脑自行设定
	'''
        self.excg_rate  = self.exchange_rate()          #设置实时人民币兑美元汇率
	self.detail_tm  = self.date_time()              #设置时间
	self.cur_dollar = 0.14516                       #人民币兑美元汇率(没网时使用)
	self.sleep_time = 2                             #写入excel的间隔时间，不能太快
	self.year_key   = 'year'                        #是否设立新表的依据
	self.all_sheet  = 'All'                         #总表的名称 
	self.inbox_dir  = '/home/shieber/automation/billstat/inbox.txt'  #暂存邮件的位置
	self.excel_dir  = '/home/shieber/automation/billstat/'	         #excel存放路径 
	self.basename   = "debt.xlsx"                   #excel表基本名
	self.cost_dict  = {
			    "1":'time',			#系统记录时间
			    "2":'rmb',			#消费金额(元)
			    "3":'dollar'}		#消费金额(美元)
	self.sheet_info = {                             #excel表项名称,时间，消费方式等等
			    "1":'Date(y-m-d h:m:s)',	#时间 
			    "2":'Method',		#消费方式
			    "3":'Expenditure(￥)',	#消费金额(元)
			    "4":'Expenditrue($)',	#消费金额(美元)
			    "5":'Notes'			#你自己添加的评注，比如买了衣服
			  }

	########1.获取人民币兑美元汇率#########################
    def exchange_rate(self):
	'''获取并返回人民币兑美元的汇率'''
	try:	
	    cur_l = self.money_search()
	except AttributeError:           
	    cur_l = [1,self.cur_dollar]                 #获取失败时返回默认值
	return cur_l

    def money_search(self):
	'''从www.currencydo.com获取匹配的字符串:汇率获取'''
	url = 'https://www.currencydo.com/'
	user_agent = 'Mozilla/4.0 (compatible; MSTE 5.5; Windows NT)'
	headers = {'User-Agent':user_agent}
	resp = requests.get(url,headers=headers)
	html = str(resp.text)

	cur_p = re.compile(r'今日1美元兑人民币汇率是：(\d\.\d+)人民币(.*)(\d\.\d+)美金')
	cur_s = cur_p.search(html)
	cur_l = [float(cur_s.group(1)),float(cur_s.group(3))]
	return cur_l

    ########2.记录系统处理时间#########################
    def date_time(self):
	'''
	    返回处理日期和时间，格式为：
	    2019-12-25 18:11:23这样的格式
	    该值是写入excel表第一列的数据
	'''
	date = datetime.datetime.now()
	year = str(date.year)
	mon  = str(date.month)
	day  = str(date.day)
	hour = str(date.hour) 
	minu = str(date.minute)
	sec  = str(date.second)
	date_s = ''.join([year,"-",mon,"-",day])
	time_s = ''.join([' ',hour,':',minu,':',sec])
	return [date_s, time_s]

    ########3.提取邮件中的消费信息######################
    def get_cost_info(self):
	'''从邮件中提取所需的消费内容'''
	emailObj = open(self.inbox_dir)
	textline = emailObj.readline()

	date_p = re.compile(r'\((\d+)(-|/|\.|\s)(\d+)(-|/|\.|\s)(\d+)\)') #正则查找消费时间
	date_s = date_p.findall(textline)
	if date_s:
	    year = str(date_s[0][0]) 
	    date = "".join(date_s[0])
	else:
	    year = str(datetime.datetime.now().year)	#邮件里没记录时间就用当前时间
	    date = self.detail_tm[0]
	textline = emailObj.readline()
	emailObj.close()

	money_p = re.compile(r'(\w+?)(:)?(\d+)(\.)?(\d+)?')   #正则查找消费数据
	money = money_p.findall(textline)
	if money:
	    cost_info = self.get_money_dic(year, date, money)
	else:
	    cost_info = {}                              #如果没有消费信息返回空字典
	return cost_info 

    def get_money_dic(self, year, date, money):
	'''
            解析消费内容成字典格式并返回
	    money格式为[('jd',':','32','.','21'),(),()]
	'''
	money_dic  = {}
	money_dic[self.year_key] = year
	for i in range(len(money)):
	    cost_tuple  = money[i]                   #解析消费数据元组
	    cost_method = cost_tuple[0]              #解析消费方式
	    money_lis   = cost_tuple[2:]             #解析消费金额字符为列表
	    money_rmb   = float(''.join(money_lis))  #拼接消费金额为小数
	    money_usd   = round(self.excg_rate[1]*money_rmb, 2) #转换为美元格式
	    sys_time    = date + self.detail_tm[1]

	    money_dic[method] = {
				  self.cost_dict['1']: sys_time,
		                  self.cost_dict['2']: money_rmb,
				  self.cost_dict['3']: money_usd
				}                    #封装所有信息为字典并返回
	return money_dic

    ########4.向Excel中写入消费信息################
    def write_to_excel(self):
	'''向20xxdebt.xlsx表中写入所有信息(核心函数)'''
	cost_info = self.get_cost_info()
	if not cost_info:
	    sys.exit(-1)                        #没有消费信息则直接退出

	excel_name = self.excel_dir + str(cost_info[self.year_key]) + self.basename 
	keys = cost_info.keys()                 #keys处容易出错
	del keys[keys.index(self.year_key)]

	if not os.path.exists(excel_name):      #判断对应年文件是否存在,不存在就创建
	    keys.append(self.all_sheet)         #加入总表All项
	    self.create_year_sheet(keys, excel_name)
	    del keys[keys.index(self.all_sheet)]#删除All项
		
	wb = openpyxl.load_workbook(excel_name)	#打开对应年的文件写入信息
	sheets = wb.get_sheet_names()
	for key in keys:
	    key_upper = key.title()		#消费方式的首字符大写
	    if key_upper not in sheets:
		wb.create_sheet(1,key_upper)	#为新消费方式添加分表
		self.add_item(wb,key_upper)

	    sheet_lis = [self.all_sheet, key_upper]                   #记录到总表和分表中
	    for sheet in sheet_lis:
                curren_s = wb.get_sheet_by_name(sheet)                #开总表和分表记录消费
                n_row  = str(curren_s.get_highest_row() + 1)          #设置写入的行数
                curren_s['A' + n_row] = cost_info[key][self.cost_dict['1']] #记录时间
                curren_s['B' + n_row] = key                                 #消费方式
                curren_s['C' + n_row] = cost_info[key][self.cost_dict['2']] #金额(人民币）
                curren_s['D' + n_row] = cost_info[key][self.cost_dict['3']] #金额(美元）
		#curren_s.cell(row=new_row, column=i).value=new_t     更健壮写入

	wb.save(excel_name)
	time.sleep(self.sleep_time)   #稍停顿，待excel表数据存储完毕，写入太快会出错

	########5.创建数据记录表20xxdebt.xlsx##########
    def create_year_sheet(self,sheets, excel_name):
	'''如果不存在某年的表就建立相应的表(20xxdebt.xlsx)'''
	wb = openpyxl.Workbook()
	for sheet in sheets:
	    wb.create_sheet(0, sheet.title())          #首字母大写
	                                         
	sheet_names = wb.get_sheet_names()
	for sheet_name in sheet_names:
	    self.add_item(wb,sheet_name)

	wb.remove_sheet(wb.get_sheet_by_name('Sheet')) #删除多余的表	
	wb.save(excel_name)

    def add_item(self,wb,name):
	'''为每个表添加消费项的标题'''
	sheet = wb.get_sheet_by_name(name)
	sheet['A1'] = self.sheet_info['1']             #初始化表的记录项的相关信息
	sheet['B1'] = self.sheet_info['2']
	sheet['C1'] = self.sheet_info['3']
	sheet['D1'] = self.sheet_info['4']
	sheet['E1'] = self.sheet_info['5']


class EmailManage():
    '''连接网络查询，下载，调用Debtmanage类函数写入excel,删除邮箱对应邮件'''
    def __init__(self):
        '''初始化设置相关信息'''
        self.debtmanage  = DebtManage()
        self.email_num   = 10                    #默认读取邮件数量 
        self.inbox_dir   = '/home/shieber/automation/bill/inbox.txt' 
        self.pop3_server = 'pop.163.com'         #对应邮箱的pop3服务器
        self.email_addr  = "your email accound"  #你的邮箱账号
        self.password    = "eamil password"      #你的登录密码
        self.identifier  = 'cost information'    #你发送的邮件标题中的关键字

    def guess_charset(self, msg): 
        '''获取邮件字符集编码'''
        charset = msg.get_charset() 
        if charset is None: 
            content_type = msg.get('Content-Type', '').lower()  #获取失败时再次获取
            pos = content_type.find('charset=') 
            if pos >= 0: 
                charset = content_type[pos + 8:].strip() 
        return charset 

    def print_info(self, msg, indent=0): 
        '''打印出邮件信息'''
        header = "Subject"
        if indent == 0: 
            value = msg.get(header, '') 
            if value: 
                print('%s: %s' % (header, value)) 
            else:
                sys.exit(-1)

        if (msg.is_multipart()): 
            parts = msg.get_payload() 
            for n, part in enumerate(parts):
                self.print_info(part, indent + 1) 
        else: 
            content_type = msg.get_content_type() 
            if content_type=='text/plain' or content_type=='text/html': 
                content = msg.get_payload(decode=True) 
                charset = self.guess_charset(msg) 
                if charset: 
                    content = content.decode(charset) 
                print('Text: %s' % (content)) 

    def connect(self):
        '''连接邮箱服务器'''
        try:
            server = poplib.POP3_SSL(self.pop3_server, 995) 
            server.user(self.email_addr) 
            server.pass_(self.password) 
        except Exception as err:
            return None, None

        resp, mails, octets = server.list() 
        indexs = len(mails) 

        return indexs, server

    def download_write(self, index, server):
        '''下载邮件并写入excel表'''
        try:
            resp, lines, octets = server.retr(index)      #下载邮件
        except Exception as err:
            sys.exit(-1)

        msg_content = b'\r\n'.join(lines).decode('utf-8') #拼接邮件内容 
        msg = Parser().parsestr(msg_content)              #解析邮件内容

        value = msg.get("Subject", '').lower()
        if self.identifier in value:	         #标题含有关键的标识符时执行
            self.write_to_inbox(msg)             #先写入暂时文件inbox.txt
            self.debtmanage.write_to_excel()     #调用DebtManage的方法,写入excel(核心函数)
            server.dele(index)                   #删除邮件,核心函数,千万不要误删

    def write_to_inbox(self, msg):
        '''写入inbox.txt'''
        with open(self.inbox_dir,'w') as inboxObj:
            stdotput = sys.stdout                    #暂时将标准输出存储起来，以便恢复   
            sys.stdout = inboxObj                    #设置系统打印输出到文件
            self.print_info(msg)                     #输出邮件信息到本地文件inbox.txt
            sys.stdout = stdotput                    #恢复系统输出到终端

    def main(self, indexs, server):
        '''下载写入并删除邮件(核心函数)'''
        if indexs > self.email_num:
            for index in range(indexs, indexs - self.email_num, -1):
                self.download_write(index, server)   #具体写入函数
        else:
            for index in range(indexs,0,-1):	     
                self.download_write(index, server) 
        server.quit() 


if __name__ == "__main__":
    email_manage = EmailManage()
    indexs, server = email_manage.connect()
    email_manage.main(indexs, server)
